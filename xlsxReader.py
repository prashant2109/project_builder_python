# -*- coding: utf-8 -*-
import sys, os
import openpyxl
from openpyxl import load_workbook
from datetime import *

class xlsxReader(object):

    def readExcel(self, fname, data_flg=False):
        excel_data = {} 
        wb = load_workbook(filename=fname, data_only=data_flg)
        sheet_names = wb.get_sheet_names()
        for idx, sheet_name in enumerate(sheet_names):
            sheetObj = wb.get_sheet_by_name(sheet_name)
            excel_data[sheet_name] = []
            for rowid, rowObjs in enumerate(sheetObj.rows):
                row = []
                for colid, cellObj in enumerate(rowObjs):
                    x = cellObj.value
                    if x == None : x = ''
                    try: x = str(x)
                    except: pass
                    x = ' '.join(map(lambda x:x.strip(), x.split()))
                    row.append(x)
                excel_data[sheet_name].append(row[:])
        return  excel_data
   
    def readExcel_withformulaflg(self, dataPath, file_name, data_flg=False):
        fname = os.path.join(dataPath, '%s.xlsx' %file_name)
        excel_data = {} 
        wb = load_workbook(filename=fname, data_only=data_flg)
        sheet_names = wb.get_sheet_names()
        for idx, sheet_name in enumerate(sheet_names):
            sheetObj = wb.get_sheet_by_name(sheet_name)
            excel_data[sheet_name] = []
            for rowid, rowObjs in enumerate(sheetObj.rows):
                row = []
                for colid, cellObj in enumerate(rowObjs):
                    x = cellObj.value
                    formula_flag = cellObj.internal_value
                    #if formula_flag:
                    #    formula_flag = 1
                    #else:
                    #    formula_flag = 0
                    if x == None : x = ''
                    try: x = str(x)
                    except: pass
                    x = ' '.join(map(lambda x:x.strip(), x.split()))
                    row.append((x, formula_flag))
                excel_data[sheet_name].append(row[:])
        return  excel_data

    def readExcel_with_comments(self, fname, data_flg=False):
        excel_data = {} 
        wb = load_workbook(filename=fname, data_only=data_flg)
        sheet_names = wb.get_sheet_names()
        for idx, sheet_name in enumerate(sheet_names):
            sheetObj = wb.get_sheet_by_name(sheet_name)
            excel_data[sheet_name] = []
            for rowid, rowObjs in enumerate(sheetObj.rows):
                row = []
                for colid, cellObj in enumerate(rowObjs):
                    x = cellObj.value
                    comment = cellObj.comment
                    data_type = cellObj.number_format
                    if x == None : x = ''
                    try: x = str(x)
                    except: pass
                    x = ' '.join(map(lambda x:x.strip(), x.split()))
                    data_tup = (x, comment, data_type)
                    row.append(data_tup)
                excel_data[sheet_name].append(row[:])
        return  excel_data


    def process(self, dataPath, file_name, data_flg=False, comment_flg=False):
        iFile = os.path.join(dataPath, '%s.xlsx' %file_name)
        if not os.path.exists(iFile): return {}
        sheet_dict = {}
        if not comment_flg:
            excel_op_data = self.readExcel(iFile, data_flg)
        elif comment_flg:
            excel_op_data = self.readExcel_with_comments(iFile, data_flg)
        return excel_op_data

#TEST
if __name__=="__main__":
    obj = xlsxReader()
    dataPath = sys.argv[1]
    caseID = sys.argv[2] 
    obj.process(dataPath, caseID)
